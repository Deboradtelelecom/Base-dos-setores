import openpyxl
wb = openpyxl.load_workbook('dados-mensais/Base_Rateio_Custos_DTEL.xlsx', data_only=True)
wsS = wb['SETORES']
wsC = wb['CUSTOS MENSAIS']
oficiais = {}
for r in wsS.iter_rows(min_row=5, values_only=True):
    nome, criterio, just, total, unit, mes = r[0], r[1], r[2], r[3], r[4], r[5]
    if not nome or not mes: continue
    oficiais[(nome, mes)] = total or 0

somas = {}
for r in wsC.iter_rows(min_row=5, values_only=True):
    setor, categoria, desc, qtd, vu, total, crit, mes = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]
    if not setor or not mes: continue
    somas[(setor, mes)] = somas.get((setor, mes), 0) + (total or 0)

problems = 0
checked = 0
for k, oficial in oficiais.items():
    soma = somas.get(k, 0)
    checked += 1
    if abs(soma - oficial) > 0.01:
        problems += 1
        print('DIFF', k, 'oficial=', oficial, 'soma=', soma, 'diff=', soma-oficial)
print('Total combos checked:', checked)
print('Problems (diff > 0.01):', problems)
